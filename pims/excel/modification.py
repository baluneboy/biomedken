#!/usr/bin/env python

import calendar
import datetime
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import Cell, get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.reader.excel import load_workbook

# OBSOLETE insert wall clock column [because pandas does it now]
def OBSOLETE_insert_wall_clock_column(ws, left_of_column=2):
    
    # inserting column to the left of input left_of_column
    column_index = left_of_column
    new_cells = {}
    ws.column_dimensions = {}
    for coordinate, cell in ws._cells.iteritems():
        column_letter, row = coordinate_from_string(coordinate)
        column = column_index_from_string(column_letter)
    
        # shifting columns
        if column >= column_index:
            column += 1
    
        column_letter = get_column_letter(column)
        coordinate = '%s%s' % (column_letter, row)
    
        # it's important to create new Cell object
        new_cells[coordinate] = Cell(ws, column_letter, row, cell.value)
        
    ws._cells = new_cells
    
    return ws

# if month GMT range okay, then overwrite last row with per-column totals
def reckon_month(ws):
    """if month GMT range okay, then overwrite last row with per-column totals"""
    
    if ws.cell('A2').value == 'Date':
        gmt_start = ws.cell('A3').value
        if gmt_start.day == 1:
            last_day = calendar.monthrange(gmt_start.year, gmt_start.month)[1]
            gmt_end = datetime.datetime(gmt_start.year, gmt_start.month, last_day)
            # go to bottom of Date column to get last GMT
            last_row = ws.get_highest_row()
            last_gmt = ws.cell('A' + str(last_row)).value
            delta_days = (last_gmt - gmt_end).days
            if delta_days == 1:
                #gmt_range_str = 'GMT range is %s through %s' % (gmt_start.strftime('%Y-%m-%d'), gmt_end.strftime('%Y-%m-%d'))
                ws.cell(row=(last_row - 1), column=0).value = 'TOTAL'
                for c in range(ws.get_highest_column())[1:]:
                    letter = get_column_letter(c + 1)
                    formula_str = "=SUM(%s3:%s%d)" % (letter, letter, (last_row - 1))
                    ws.cell(row=(last_row - 1), column=c).value = formula_str
            else:
                print 'Abort: last_gmt = %s is not <= 2 days delta from end of month = %s' % (last_gmt, gmt_end)
        else:
            print 'gmt_start is not day one'
    else:
        print 'A2 is not Date'
        
    return ws, gmt_start, gmt_end

# reckon GMT range and overwrite last row with totals, if okay
def overwrite_last_row_with_totals(xlsxfile):
    """reckon GMT range and overwrite last row with totals, if okay"""
    
    # load workbook and get "raw" worksheet and "kpi" worksheet
    wb = load_workbook(filename = xlsxfile)
    ws = wb.get_sheet_by_name("raw")
    
    # check GMT range for month
    ws, gmt_start, gmt_end = reckon_month(ws)
    gmt_range_str = 'GMT range is %s through %s' % (gmt_start.strftime('%Y-%m-%d'), gmt_end.strftime('%Y-%m-%d'))
    ws.cell('A1').value = '/</-/'
    
    # save
    wb.save(filename = xlsxfile)
    
    return gmt_start, gmt_end

def demo_create_write():
    wb = Workbook()
    
    dest_filename = r'/tmp/empty_book.xlsx'
    
    ws = wb.worksheets[0]
    
    ws.title = "range names"
    
    for col_idx in xrange(1, 40):
        col = get_column_letter(col_idx)
        for row in xrange(1, 600):
            ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)
    
    ws = wb.create_sheet()
    
    ws.title = 'Pi'
    
    ws.cell('F5').value = 3.14
    
    wb.save(filename = dest_filename)
        
def update_xlsx(src, dest):
    #Open an xlsx for reading
    wb = load_workbook(filename = src)
    
    ##Get the current Active Sheet
    #ws = wb.get_active_sheet()
    
    #You can also select a particular sheet name
    ws = wb.get_sheet_by_name("ER3grp")
    ws.cell(row=0, column=0).delete()
    
    ##Open the csv file
    #with open(src) as fin:
    #    #read the csv
    #    reader = csv.reader(fin)
    #    #enumerate the rows, so that you can
    #    #get the row index for the xlsx
    #    for index,row in enumerate(reader):
    #        #Asssuming space sepeated,
    #        #Split the row to cells (column)
    #        row = row[0].split()
    #        #Access the particular cell and assign
    #        #the value from the csv row
    #        ws.cell(row=index,column=7).value = row[2]
    #        ws.cell(row=index,column=8).value = row[3]
            
    #save the file
    wb.save(dest)
    
if __name__ == "__main__":
    #update_xlsx('/tmp/empty_book.xlsx', '/tmp/new_empty.xlsx')
    demo_open_existing()